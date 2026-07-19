from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from contexts.template.domain.template import Template

# Example monthly suffix shown on dynamic columns in the skeleton.
EXAMPLE_MONTH = "2026-01"

_INVALID_SHEET_CHARS = re.compile(r'[\\/*?\[\]:]')
_MAX_SHEET_TITLE = 31
_COLUMN_WIDTH = 20


def sheet_name_for(template: Template) -> str:
    """Sheet name: the sheet_pattern with glob wildcards stripped."""
    name = _INVALID_SHEET_CHARS.sub("", template.sheet_pattern or "").strip()
    if not name:
        name = str(template.id)
    return name[:_MAX_SHEET_TITLE]


def build_template_workbook(template: Template) -> bytes:
    """Render an .xlsx skeleton for a template: sheet named after the sheet
    pattern, header rows laid out per ``headers.rows``, one column per fixed
    column (first match keyword as the column name), example monthly columns
    for dynamic columns, and blank data rows from ``data_start_row``."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name_for(template)

    header_rows = max(len(template.header_spec.header_rows), 1)
    last_header_row = header_rows

    col = 1
    if template.hierarchy_config is not None and template.hierarchy_config.column_name:
        ws.cell(
            row=last_header_row, column=col,
            value=template.hierarchy_config.column_name.strip(),
        )
        col += 1

    seen_match_headers: set[tuple[str, ...]] = set()
    for column in template.fixed_columns:
        key = tuple(column.match_headers)
        if key in seen_match_headers:
            continue
        seen_match_headers.add(key)
        keywords = list(column.match_headers) or [column.db_field]
        # Multi-row headers: keywords run top-down and are bottom-aligned
        # (single keyword sits on the last header row, next to the data);
        # any overflow is joined into the last header row.
        if len(keywords) <= header_rows:
            start_row = last_header_row - len(keywords) + 1
            for i, keyword in enumerate(keywords):
                ws.cell(row=start_row + i, column=col, value=keyword)
        else:
            for i, keyword in enumerate(keywords[: header_rows - 1]):
                ws.cell(row=i + 1, column=col, value=keyword)
            ws.cell(
                row=last_header_row, column=col,
                value="_".join(keywords[header_rows - 1:]),
            )
        col += 1

    for dyn in template.dynamic_columns:
        ws.cell(
            row=last_header_row, column=col,
            value=f"{dyn.db_prefix}_{EXAMPLE_MONTH}",
        )
        col += 1

    for c in range(1, col):
        ws.column_dimensions[get_column_letter(c)].width = _COLUMN_WIDTH

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
