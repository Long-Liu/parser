"""XLSX export builders for analytics reports.

Pure workbook construction with openpyxl: no ORM, no HTTP. The interface
layer composes these builders with repository read models and attaches an
RFC 5987 Content-Disposition (ASCII fallback + percent-encoded UTF-8 name).
"""

from __future__ import annotations

from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_COLUMN_WIDTH = 18

# 四口径顺序与 _profit_item 一致：投标 / 指标(考核) / 当前(实际) / 预计完工
_PROFIT_CALIBERS = (
    ("bid", "投标"),
    ("indicator", "指标"),
    ("current", "当前"),
    ("forecast", "预计完工"),
)

# 多项目对比 9 项核心指标（与 compare_projects 输出及 UI 对比表一致）
_COMPARE_METRICS = (
    ("progress", "进度(%)"),
    ("contract", "合同价"),
    ("settlement", "结算产值"),
    ("revenue", "营收"),
    ("total_cost", "总成本"),
    ("profit", "毛利"),
    ("profit_rate", "毛利率(%)"),
    ("settlement_rate", "结算完成率(%)"),
    ("revenue_ratio", "营收比率(%)"),
)

_SCORE_DIMENSIONS = (
    ("profitability", "盈利能力评分"),
    ("cost_control", "成本管控评分"),
    ("progress_execution", "进度执行评分"),
    ("settlement_quality", "结算质量评分"),
    ("revenue_conversion", "营收转化评分"),
)

# 月度对比导出指标（mom 环比口径见 analytics_repository._mom_change）
_MONTH_METRICS = (
    ("revenue", "收入"),
    ("cost", "成本"),
    ("profit", "毛利"),
    ("profit_rate", "毛利率(%)"),
)


def content_disposition(filename: str, fallback: str) -> str:
    """RFC 5987: ASCII fallback filename + percent-encoded UTF-8 filename*."""
    return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{quote(filename)}"


def _autosize(ws, width: int = _COLUMN_WIDTH) -> None:
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_profits_workbook(items: list[dict]) -> Workbook:
    """项目毛利导出：四口径 × (revenue/cost/profit/profit_rate) 全列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "项目毛利"
    header = ["项目编号", "项目名称", "月份"]
    for _, label in _PROFIT_CALIBERS:
        header += [f"{label}收入", f"{label}成本", f"{label}毛利", f"{label}毛利率(%)"]
    ws.append(header)
    for item in items:
        row = [item["project_code"], item["project_name"], item["ym"]]
        for key, _ in _PROFIT_CALIBERS:
            group = item[key]
            row += [group["revenue"], group["cost"], group["profit"], group["profit_rate"]]
        ws.append(row)
    _autosize(ws)
    return wb


def _append_cost_rows(ws, projects: list[dict]) -> None:
    for proj in projects:
        for item in proj["items"]:
            ws.append([
                proj["project"]["name"], proj["ym"], item["name"],
                item["indicator"], item["actual"], item["deviation"],
                item["deviation_rate"], item["list_target"], item["adj_target"],
                item["budget"], item["forecast"],
            ])


_COST_HEADER = ["项目", "月份", "科目", "指标", "实际", "偏差", "偏差率(%)",
                "预计完工量含税指标", "调整后指标", "现执行预算", "预计完工成本"]


def build_cost_categories_workbook(projects: list[dict]) -> Workbook:
    """成本科目导出：指标/实际/偏差/偏差率 + 六口径补充列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成本科目"
    ws.append(_COST_HEADER)
    _append_cost_rows(ws, projects)
    _autosize(ws)
    return wb


def build_month_comparison_workbook(data: dict) -> Workbook:
    """月度对比导出：指标 × 月份 + 最近两月环比（变化量/变化率）列。"""
    months = data["months"]
    wb = Workbook()
    ws = wb.active
    ws.title = "月度对比"
    ws.append(["指标"] + [m["ym"] for m in months] + ["环比变化", "环比变化率(%)"])
    last_mom = months[-1]["mom"] if months else None
    for key, label in _MONTH_METRICS:
        row = [label] + [m[key] for m in months]
        if last_mom is None:
            row += [None, None]
        else:
            row += [last_mom[key]["change"], last_mom[key]["change_pct"]]
        ws.append(row)
    _autosize(ws)
    return wb


def build_compare_workbook(data: dict) -> Workbook:
    """多项目对比导出：9 项指标 × 项目 + 五维评分/等级，附成本科目 sheet。"""
    projects = data["projects"]
    wb = Workbook()
    ws = wb.active
    ws.title = "指标对比"
    ws.append(["指标"] + [p["project_name"] for p in projects])
    for key, label in _COMPARE_METRICS:
        ws.append([label] + [p[key] for p in projects])
    for key, label in _SCORE_DIMENSIONS:
        ws.append([label] + [p["scores"][key] for p in projects])
    ws.append(["综合评分"] + [p["total_score"] for p in projects])
    ws.append(["综合评级"] + [p["grade"] for p in projects])
    _autosize(ws)
    costs = wb.create_sheet("成本科目")
    costs.append(_COST_HEADER)
    _append_cost_rows(costs, data["cost_categories"])
    _autosize(costs)
    return wb
