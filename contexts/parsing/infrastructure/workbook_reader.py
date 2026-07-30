from __future__ import annotations

import asyncio
from contextlib import closing

import openpyxl

from contexts.parsing.domain.cell_unmerger import MergedCellRange
from contexts.parsing.domain.workbook import WorkbookReader, WorkbookSheet


def worksheet_to_grid(ws) -> tuple[list[list], list[MergedCellRange]]:
    ranges = [
        MergedCellRange(
            min_col=merged_range.min_col - 1,
            max_col=merged_range.max_col - 1,
            min_row=merged_range.min_row - 1,
            max_row=merged_range.max_row - 1,
        )
        for merged_range in ws.merged_cells.ranges
    ]
    # Excel's used-range metadata is frequently inflated by formatting.  Some
    # real uploads report 65k rows or all 16,384 columns for only a handful of
    # populated cells.  Determine bounds from cells that actually carry values,
    # while retaining merged-range edges required by CellUnmerger.
    populated = [cell for cell in ws._cells.values() if cell.value is not None]
    max_row = max(
        [cell.row for cell in populated] + [item.max_row + 1 for item in ranges],
        default=0,
    )
    max_col = max(
        [cell.column for cell in populated] + [item.max_col + 1 for item in ranges],
        default=0,
    )
    if max_row == 0 or max_col == 0:
        return [], ranges
    grid = [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
    ]
    return grid, ranges


class OpenPyxlWorkbookReader(WorkbookReader):
    async def read(self, filepath: str) -> list[WorkbookSheet]:
        return await asyncio.to_thread(self._read_sync, filepath)

    @staticmethod
    def _read_sync(filepath: str) -> list[WorkbookSheet]:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        with closing(wb):
            sheets = []
            for sheet_name in wb.sheetnames:
                grid, ranges = worksheet_to_grid(wb[sheet_name])
                sheets.append(WorkbookSheet(name=sheet_name, grid=grid, merged_ranges=ranges))
            return sheets
