import openpyxl
from core.pipeline import run_pipeline


def make_test_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表1 人工费-动态"
    # Row1 (header_row): 序号 | 姓名 | 部门  (0-indexed: row 0)
    ws["A1"] = "序号"; ws["B1"] = "姓名"; ws["C1"] = "部门"
    # Row2: data
    ws["A2"] = "1.1"; ws["B2"] = "张三"; ws["C2"] = "技术部"
    ws["A3"] = "1.2"; ws["B3"] = "李四"; ws["C3"] = "经营部"
    return wb


def make_config():
    return {
        "template_id": "test_labor",
        "sheet_pattern": "表1*",
        "headers": {"rows": [1], "data_start_row": 2},
        "hierarchy": {"column_name": "序号", "separator": "."},
        "columns": [
            {"db_field": "person_name", "match_header": ["姓名"], "type": "varchar(100)"},
            {"db_field": "dept", "match_header": ["部门"], "type": "varchar(100)"},
        ],
        "dynamic_columns": [],
        "stop_rules": [
            {"type": "cell_match", "patterns": ["^注："], "columns": ["A"]},
        ],
    }


def test_pipeline_extracts_data():
    wb = make_test_workbook()
    ws = wb.active
    config = make_config()
    result = run_pipeline(ws, batch_id=1, config=config)

    assert result["template_id"] == "test_labor"
    assert result["total_rows"] == 2
    assert result["success_rows"] == 2
    assert len(result["rows"]) == 2
    assert result["rows"][0]["person_name"] == "张三"
    assert result["rows"][0]["hierarchy_code"] == "1.1"


def test_pipeline_stops_on_comment_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表1 测试"
    # Row1: header  (data_start_row=2, so row1 is header, row2+ is data)
    ws["A1"] = "序号"; ws["B1"] = "姓名"
    # Row2: data
    ws["A2"] = "1"; ws["B2"] = "数据1"
    # Row3: comment - should stop
    ws["A3"] = "注：以下为说明"
    # Row4: should not be read
    ws["A4"] = "不应该被读取"; ws["B4"] = "忽略"

    config = {
        "template_id": "test",
        "sheet_pattern": "表1*",
        "headers": {"rows": [1], "data_start_row": 2},
        "hierarchy": {"column_name": "序号", "separator": "."},
        "columns": [
            {"db_field": "person_name", "match_header": ["姓名"], "type": "varchar(100)"},
        ],
        "dynamic_columns": [],
        "stop_rules": [
            {"type": "cell_match", "patterns": ["^注："], "columns": ["A"]},
        ],
    }

    result = run_pipeline(ws, batch_id=1, config=config)

    assert result["total_rows"] == 1
    assert result["rows"][0]["person_name"] == "数据1"
