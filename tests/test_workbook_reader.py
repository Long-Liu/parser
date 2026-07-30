from copy import copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from contexts.parsing.infrastructure.workbook_reader import worksheet_to_grid


def test_worksheet_to_grid_ignores_format_only_used_range():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "名称"
    ws["A2"] = "数据"
    ws.merge_cells("A3:C3")
    ws["A3"] = "合计"
    ws["XFD8"].fill = copy(PatternFill(fill_type="solid", fgColor="FFFF00"))

    grid, ranges = worksheet_to_grid(ws)

    assert len(grid) == 3
    assert len(grid[0]) == 3
    assert grid[2][0] == "合计"
    assert len(ranges) == 1
