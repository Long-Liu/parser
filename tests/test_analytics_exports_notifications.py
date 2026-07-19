"""Tests for analytics exports, notification management, and compare AI report.

Covers: four-caliber profit export columns + RFC 5987 Chinese filenames + the
removed 100-row cap, six-caliber cost-category export, month-comparison and
multi-project compare export workbooks, notification read-all/delete/clear
(including cross-user protection), and the deterministic five-chapter
multi-project AI report fallback.

Profit figures come from 表11 结算产值表 (data_settlement_output): the sheet
has no bid/target-indicator split, so the bid/indicator calibers report zeros
while current/forecast carry the settlement values.

Controller handlers are invoked via ``__wrapped__`` unwrapping to exercise the
handler body without the auth decorators; the 401/403 decorator contract
itself is covered by tests/test_auth_middleware.py and the endpoint smoke
suite.
"""

import io
import json as jsonlib
from decimal import Decimal
from itertools import count
from types import SimpleNamespace
from urllib.parse import quote

import pytest
from openpyxl import load_workbook
from tortoise import Tortoise

from contexts.analytics.application.analytics_service import (
    AnalyticsApplicationService,
)
from contexts.analytics.domain.compare_report import build_compare_report
from contexts.analytics.infrastructure.analytics_repository import (
    TortoiseAnalyticsRepository,
)
from contexts.analytics.infrastructure.xlsx_export import (
    build_compare_workbook,
    build_cost_categories_workbook,
    build_month_comparison_workbook,
    build_profits_workbook,
    content_disposition,
)
from contexts.analytics.interface.analytics_controller import AnalyticsController
from contexts.auth.infrastructure.tables import Notification, NotificationRead
from contexts.parsing.infrastructure.tables import UploadBatch
from contexts.project.infrastructure.tables import Project
from contexts.shared.domain.exceptions import (
    AuthorizationError,
    NotFoundError,
    ValidationError,
)
from contexts.shared.domain.pagination import Pagination
from contexts.shared.infrastructure.database.engine import _MODEL_MODULES
from contexts.shared.infrastructure.database.tables import (
    SETTLE_CONTRACT_PRICE,
    SETTLE_CUMULATIVE_COST,
    SETTLE_CUMULATIVE_OUTPUT,
    SETTLE_CURRENT_PROFIT,
    SETTLE_FORECAST_COST,
    SETTLE_FORECAST_PROFIT,
    SETTLE_FORECAST_REVENUE,
    DataDynamicIndicator,
    DataSettlementOutput,
)


@pytest.fixture
async def db():
    await Tortoise.init(
        db_url="sqlite://:memory:",
        modules={"models": list(_MODEL_MODULES)},
    )
    await Tortoise.generate_schemas()
    yield
    await Tortoise.close_connections()


_seq = count(1)


async def make_project(**kwargs) -> Project:
    n = next(_seq)
    defaults = {
        "code": f"P{n:04d}", "name": f"项目{n}",
        "contract_price": Decimal("1000"), "progress": Decimal("80"),
        "status": "normal",
    }
    defaults.update(kwargs)
    return await Project.create(**defaults)


async def make_batch(project_id: int, ym: str) -> UploadBatch:
    return await UploadBatch.create(
        batch_no=f"T{next(_seq):06d}", project_id=project_id, ym=ym,
        file_name="cost.xlsx", status="success",
    )


async def make_settlement(batch_id: int, **indicators) -> None:
    """Create 表11 settlement rows: one vertical row per indicator name."""
    for name, value in indicators.items():
        await DataSettlementOutput.create(
            batch_id=batch_id, indicator_name=name,
            cumulative_value=Decimal(str(value)),
        )


async def make_profit_with_calibers(batch_id: int) -> None:
    """Settlement rows covering the current + forecast profit calibers.

    The settlement sheet has no bid / target-indicator split, so those two
    calibers report zeros by design (see repository _profit_item).
    """
    await make_settlement(
        batch_id,
        **{
            SETTLE_CONTRACT_PRICE: "1000",
            SETTLE_CUMULATIVE_OUTPUT: "900",
            SETTLE_CUMULATIVE_COST: "810",
            SETTLE_CURRENT_PROFIT: "90",
            SETTLE_FORECAST_REVENUE: "1050",
            SETTLE_FORECAST_COST: "920",
            SETTLE_FORECAST_PROFIT: "130",
        }
    )


def _controller(repo=None, access_policy=None) -> AnalyticsController:
    repo = repo or TortoiseAnalyticsRepository()
    svc = AnalyticsApplicationService(repo)
    controller = AnalyticsController(svc, access_policy, alert_svc=None)
    controller.setup()
    return controller


def _request(args=None, body=None, user_id=1, permissions=("user:manage",)):
    return SimpleNamespace(
        args=args or {},
        json=body,
        ctx=SimpleNamespace(user_id=user_id, permissions=set(permissions)),
    )


# ── xlsx export builders ─────────────────────────────────────────────


def test_content_disposition_rfc5987():
    disposition = content_disposition("项目毛利情况_2026-03.xlsx", "project-profits.xlsx")
    assert disposition.startswith('attachment; filename="project-profits.xlsx"')
    assert f"filename*=UTF-8''{quote('项目毛利情况_2026-03.xlsx')}" in disposition


@pytest.mark.asyncio
async def test_profits_workbook_has_four_calibers(db):
    project = await make_project()
    batch = await make_batch(project.id, "2026-03")
    await make_profit_with_calibers(batch.id)

    repo = TortoiseAnalyticsRepository()
    result = await repo.project_profits("2026-03", Pagination(1, 10000, max_size=10000))
    wb = build_profits_workbook(result["projects"])
    ws = wb.active

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == [
        "项目编号", "项目名称", "月份",
        "投标收入", "投标成本", "投标毛利", "投标毛利率(%)",
        "指标收入", "指标成本", "指标毛利", "指标毛利率(%)",
        "当前收入", "当前成本", "当前毛利", "当前毛利率(%)",
        "预计完工收入", "预计完工成本", "预计完工毛利", "预计完工毛利率(%)",
    ]
    row = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert row[0] == project.code and row[2] == "2026-03"
    assert row[3:7] == [0.0, 0.0, 0.0, 0.0]                 # bid: no source -> zeros
    assert row[7:11] == [0.0, 0.0, 0.0, 0.0]                # indicator: no source
    assert row[11:15] == [900.0, 810.0, 90.0, 10.0]         # current
    assert row[15:19] == [1050.0, 920.0, 130.0, 12.38]      # forecast


@pytest.mark.asyncio
async def test_cost_categories_workbook_has_six_calibers(db):
    project = await make_project()
    batch = await make_batch(project.id, "2026-03")
    await DataDynamicIndicator.create(
        batch_id=batch.id, hierarchy_code="1", item_name="安装工程",
        indicator_with_tax=Decimal("2230"), estimated_with_tax=Decimal("2183"),
        adjusted_with_tax=Decimal("2108"), current_budget=Decimal("2032"),
        incurred_cost=Decimal("1710"),
    )

    repo = TortoiseAnalyticsRepository()
    result = await repo.cost_categories(
        [project.id], "2026-03", Pagination(1, 10000, max_size=10000)
    )
    ws = build_cost_categories_workbook(result["projects"]).active

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == ["项目", "月份", "科目", "指标", "实际", "偏差", "偏差率(%)",
                      "预计完工量含税指标", "调整后指标", "现执行预算", "预计完工成本"]
    row = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert row[2] == "安装工程"
    assert row[7:11] == [2183.0, 2108.0, 2032.0, 2183.0]


@pytest.mark.asyncio
async def test_month_comparison_workbook_metrics_and_mom(db):
    project = await make_project()
    for ym, revenue, cost, net in (
            ("2026-02", "100", "90", "10"), ("2026-03", "150", "120", "30")):
        batch = await make_batch(project.id, ym)
        await make_settlement(
            batch.id,
            **{
                SETTLE_CUMULATIVE_OUTPUT: revenue,
                SETTLE_CUMULATIVE_COST: cost,
                SETTLE_CURRENT_PROFIT: net,
            }
        )

    repo = TortoiseAnalyticsRepository()
    result = await repo.month_comparison(project.id, ["2026-02", "2026-03"])
    ws = build_month_comparison_workbook(result).active

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == ["指标", "2026-02", "2026-03", "环比变化", "环比变化率(%)"]
    revenue_row = [ws.cell(row=2, column=c).value for c in range(1, 6)]
    assert revenue_row == ["收入", 100.0, 150.0, 50.0, 50.0]
    rate_row = [ws.cell(row=5, column=c).value for c in range(1, 6)]
    assert rate_row[0] == "毛利率(%)" and rate_row[3] == 10.0  # 毛利率环比以 pp 表示


@pytest.mark.asyncio
async def test_compare_workbook_metrics_scores_and_cost_sheet(db):
    alpha = await make_project(contract_price=Decimal("12500"), progress=Decimal("82"))
    batch_a = await make_batch(alpha.id, "2026-03")
    await make_profit_with_calibers(batch_a.id)
    await DataDynamicIndicator.create(
        batch_id=batch_a.id, item_name="安装工程",
        indicator_with_tax=Decimal("100"), incurred_cost=Decimal("90"),
    )
    beta = await make_project(contract_price=Decimal("8800"), progress=Decimal("75"))
    await make_batch(beta.id, "2026-03")

    repo = TortoiseAnalyticsRepository()
    result = await repo.compare_projects([alpha.id, beta.id], "2026-03")
    wb = build_compare_workbook(result)

    ws = wb["指标对比"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == ["指标", alpha.name, beta.name]
    labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert labels == [
        "进度(%)", "合同价", "结算产值", "营收", "总成本", "毛利",
        "毛利率(%)", "结算完成率(%)", "营收比率(%)",
        "盈利能力评分", "成本管控评分", "进度执行评分", "结算质量评分", "营收转化评分",
        "综合评分", "综合评级",
    ]
    grade_row = [ws.cell(row=ws.max_row, column=c).value for c in range(1, 4)]
    assert grade_row[1] in {"A", "B", "C", "D"} and grade_row[2] in {"A", "B", "C", "D"}

    costs = wb["成本科目"]
    assert costs.cell(row=2, column=3).value == "安装工程"
    # 六口径列齐全
    assert costs.max_column == 11


# ── export endpoints (unwrapped handlers) ────────────────────────────


@pytest.mark.asyncio
async def test_export_profits_endpoint_chinese_filename_and_no_100_cap(db):
    for _ in range(105):
        project = await make_project()
        batch = await make_batch(project.id, "2026-03")
        await make_profit_with_calibers(batch.id)

    controller = _controller()
    raw = AnalyticsController.export_profits.__wrapped__.__wrapped__
    response = await raw(controller, _request(args={"ym": "2026-03"}))

    assert response.status == 200
    disposition = response.headers["Content-Disposition"]
    assert 'filename="project-profits.xlsx"' in disposition
    assert f"filename*=UTF-8''{quote('项目毛利情况_2026-03.xlsx')}" in disposition

    ws = load_workbook(io.BytesIO(response.body)).active
    assert ws.max_row == 1 + 105  # 超出原硬编码 100 条限制
    assert ws.max_column == 19


@pytest.mark.asyncio
async def test_export_costs_endpoint_chinese_filename(db):
    project = await make_project()
    batch = await make_batch(project.id, "2026-03")
    await DataDynamicIndicator.create(
        batch_id=batch.id, item_name="安装工程",
        indicator_with_tax=Decimal("100"), incurred_cost=Decimal("90"),
    )

    controller = _controller()
    raw = AnalyticsController.export_costs.__wrapped__.__wrapped__
    response = await raw(controller, _request(
        args={"ym": "2026-03", "project_ids": str(project.id)}
    ))

    disposition = response.headers["Content-Disposition"]
    assert 'filename="cost-categories.xlsx"' in disposition
    assert f"filename*=UTF-8''{quote('成本科目_2026-03.xlsx')}" in disposition
    ws = load_workbook(io.BytesIO(response.body)).active
    assert ws.max_row == 2 and ws.max_column == 11


@pytest.mark.asyncio
async def test_export_month_comparison_endpoint(db):
    project = await make_project()
    for ym in ("2026-02", "2026-03"):
        batch = await make_batch(project.id, ym)
        await make_settlement(
            batch.id,
            **{
                SETTLE_CUMULATIVE_OUTPUT: "100",
                SETTLE_CUMULATIVE_COST: "90",
                SETTLE_CURRENT_PROFIT: "10",
            }
        )

    controller = _controller()
    raw = (AnalyticsController.export_month_comparison
           .__wrapped__.__wrapped__.__wrapped__)
    response = await raw(controller, _request(
        args={"months": "2026-02,2026-03"}
    ), project.id)

    disposition = response.headers["Content-Disposition"]
    assert f"filename*=UTF-8''{quote('月度对比_2026-02_2026-03.xlsx')}" in disposition
    ws = load_workbook(io.BytesIO(response.body)).active
    assert [ws.cell(row=1, column=c).value for c in range(1, 6)] == [
        "指标", "2026-02", "2026-03", "环比变化", "环比变化率(%)",
    ]


@pytest.mark.asyncio
async def test_export_compare_endpoint(db):
    alpha = await make_project()
    await make_batch(alpha.id, "2026-03")
    beta = await make_project()
    await make_batch(beta.id, "2026-03")

    controller = _controller()
    raw = AnalyticsController.export_compare.__wrapped__.__wrapped__
    response = await raw(controller, _request(
        args={"project_ids": f"{alpha.id},{beta.id}", "ym": "2026-03"}
    ))

    disposition = response.headers["Content-Disposition"]
    assert f"filename*=UTF-8''{quote('多项目对比_2026-03.xlsx')}" in disposition
    wb = load_workbook(io.BytesIO(response.body))
    assert wb.sheetnames == ["指标对比", "成本科目"]


@pytest.mark.asyncio
async def test_project_profits_report_endpoint_still_works(db):
    """Regression: the GET report handler passed page/size instead of Pagination."""
    project = await make_project()
    batch = await make_batch(project.id, "2026-03")
    await make_profit_with_calibers(batch.id)

    controller = _controller()
    raw = AnalyticsController.project_profits.__wrapped__.__wrapped__
    response = await raw(controller, _request(args={"ym": "2026-03"}))

    payload = jsonlib.loads(response.body)
    assert payload["pagination"]["total"] == 1
    item = payload["projects"][0]
    assert set(item) >= {"bid", "indicator", "current", "forecast"}
    # settlement sheet has no bid split -> zeros; current carries real values
    assert item["bid"]["profit"] == 0.0
    assert item["current"]["profit"] == 90.0
    assert item["current"]["profit_rate"] == 10.0


# ── notification repository ──────────────────────────────────────────


async def make_notification(user_id, title="通知") -> Notification:
    return await Notification.create(
        user_id=user_id, notification_type="system",
        title=title, message="内容",
    )


@pytest.mark.asyncio
async def test_mark_all_notifications_read_own_and_broadcast(db):
    own = await make_notification(1, "本人")
    broadcast = await make_notification(None, "广播")
    other = await make_notification(2, "他人")

    repo = TortoiseAnalyticsRepository()
    marked = await repo.mark_all_notifications_read(1)
    assert marked == 2  # 本人 + 广播；他人通知不可见

    again = await repo.mark_all_notifications_read(1)
    assert again == 0  # 幂等

    read_ids = set(await NotificationRead.filter(user_id=1).values_list(
        "notification_id", flat=True))
    assert read_ids == {own.id, broadcast.id}
    assert not await NotificationRead.filter(
        user_id=1, notification_id=other.id).exists()


@pytest.mark.asyncio
async def test_delete_notification_only_own(db):
    mine = await make_notification(1)
    other = await make_notification(2)
    broadcast = await make_notification(None)

    repo = TortoiseAnalyticsRepository()
    await repo.delete_notification(1, mine.id)
    assert await Notification.get_or_none(id=mine.id) is None

    # 他人通知、广播通知均不可删 -> 404 语义
    with pytest.raises(NotFoundError):
        await repo.delete_notification(1, other.id)
    with pytest.raises(NotFoundError):
        await repo.delete_notification(1, broadcast.id)
    assert await Notification.get_or_none(id=other.id) is not None
    assert await Notification.get_or_none(id=broadcast.id) is not None


@pytest.mark.asyncio
async def test_clear_notifications_keeps_broadcast_and_others(db):
    await make_notification(1, "本人1")
    await make_notification(1, "本人2")
    other = await make_notification(2)
    broadcast = await make_notification(None)

    repo = TortoiseAnalyticsRepository()
    deleted = await repo.clear_notifications(1)
    assert deleted == 2
    remaining = await Notification.all()
    assert {n.id for n in remaining} == {other.id, broadcast.id}

    assert await repo.clear_notifications(1) == 0


# ── notification endpoints (unwrapped handlers) ──────────────────────


@pytest.mark.asyncio
async def test_notification_endpoints_via_controller(db):
    mine = await make_notification(1)
    broadcast = await make_notification(None)

    controller = _controller()
    mark_all = AnalyticsController.mark_all_read.__wrapped__
    response = await mark_all(controller, _request(user_id=1))
    assert jsonlib.loads(response.body) == {"ok": True, "marked": 2}

    delete_one = AnalyticsController.delete_notification.__wrapped__
    response = await delete_one(controller, _request(user_id=1), mine.id)
    assert jsonlib.loads(response.body) == {"ok": True}

    clear = AnalyticsController.clear_notifications.__wrapped__
    response = await clear(controller, _request(user_id=1))
    assert jsonlib.loads(response.body) == {"ok": True, "deleted": 0}
    assert await Notification.get_or_none(id=broadcast.id) is not None


@pytest.mark.asyncio
async def test_delete_notification_endpoint_rejects_others_notification(db):
    other = await make_notification(2)
    controller = _controller()
    delete_one = AnalyticsController.delete_notification.__wrapped__
    with pytest.raises(NotFoundError):
        await delete_one(controller, _request(user_id=1), other.id)


# ── multi-project AI report ──────────────────────────────────────────


async def _seed_compare_projects():
    alpha = await make_project(contract_price=Decimal("12500"), progress=Decimal("82"))
    batch_a = await make_batch(alpha.id, "2026-03")
    await make_settlement(
        batch_a.id,
        **{
            SETTLE_CUMULATIVE_OUTPUT: "10250",
            SETTLE_CUMULATIVE_COST: "8050",
            SETTLE_CURRENT_PROFIT: "2200",
        }
    )
    beta = await make_project(contract_price=Decimal("8800"), progress=Decimal("75"))
    batch_b = await make_batch(beta.id, "2026-03")
    await make_settlement(
        batch_b.id,
        **{
            SETTLE_CUMULATIVE_OUTPUT: "6600",
            SETTLE_CUMULATIVE_COST: "5060",
            SETTLE_CURRENT_PROFIT: "1540",
        }
    )
    return alpha, beta


@pytest.mark.asyncio
async def test_compare_ai_analysis_fallback_chapters(db):
    alpha, beta = await _seed_compare_projects()

    repo = TortoiseAnalyticsRepository()  # no provider -> deterministic fallback
    result = await repo.compare_ai_analysis([alpha.id, beta.id], "2026-03")

    assert result["ym"] == "2026-03"
    assert result["project_ids"] == [alpha.id, beta.id]
    assert result["generated_at"]  # 服务端生成时间，非硬编码
    chapters = result["chapters"]
    assert [c["key"] for c in chapters] == [
        "overview", "progress", "cost", "profit", "rating",
    ]
    assert [c["title"] for c in chapters] == [
        "核心经营总览", "全项目进度对标", "成本专项经营分析",
        "盈利专项分析", "项目综合评级",
    ]
    assert all(c["content"] for c in chapters)
    assert {p["grade"] for p in result["projects"]} == {"A", "B"}
    # 最优项目（alpha）应出现在总览文案中
    assert alpha.name in chapters[0]["content"]


class _FakeProvider:
    def __init__(self, result):
        self._result = result
        self.payload = None

    async def analyze(self, payload):
        self.payload = payload
        return self._result


@pytest.mark.asyncio
async def test_compare_ai_analysis_uses_provider_when_available(db):
    alpha, beta = await _seed_compare_projects()
    provider = _FakeProvider({"chapters": [{"key": "x", "title": "t", "content": "c"}]})

    repo = TortoiseAnalyticsRepository(provider)
    result = await repo.compare_ai_analysis([alpha.id, beta.id], "2026-03")

    assert result["chapters"] == [{"key": "x", "title": "t", "content": "c"}]
    assert result["generated_at"]
    assert provider.payload["type"] == "project_comparison"
    assert provider.payload["period"] == "2026-03"
    assert len(provider.payload["projects"]) == 2


@pytest.mark.asyncio
async def test_compare_ai_analysis_falls_back_when_provider_returns_none(db):
    alpha, beta = await _seed_compare_projects()
    repo = TortoiseAnalyticsRepository(_FakeProvider(None))
    result = await repo.compare_ai_analysis([alpha.id, beta.id], "2026-03")
    assert [c["key"] for c in result["chapters"]] == [
        "overview", "progress", "cost", "profit", "rating",
    ]


@pytest.mark.asyncio
async def test_compare_ai_analysis_endpoint_validates_ids(db):
    controller = _controller()
    raw = AnalyticsController.compare_ai_analysis.__wrapped__.__wrapped__
    with pytest.raises(ValidationError, match="invalid project_ids"):
        await raw(controller, _request(body={"project_ids": ["abc"]}))


class _DenyAccessPolicy:
    async def accessible_project_ids(self, user_id):
        return {1}  # 仅项目 1 可见


@pytest.mark.asyncio
async def test_compare_ai_analysis_endpoint_enforces_project_scope(db):
    await _seed_compare_projects()
    controller = _controller(access_policy=_DenyAccessPolicy())
    raw = AnalyticsController.compare_ai_analysis.__wrapped__.__wrapped__
    with pytest.raises(AuthorizationError):
        await raw(controller, _request(
            body={"project_ids": [1, 2], "ym": "2026-03"},
            permissions=("data:view",),
        ))


@pytest.mark.asyncio
async def test_compare_ai_analysis_endpoint_returns_report(db):
    alpha, beta = await _seed_compare_projects()
    controller = _controller()
    raw = AnalyticsController.compare_ai_analysis.__wrapped__.__wrapped__
    response = await raw(controller, _request(
        body={"project_ids": [alpha.id, beta.id], "ym": "2026-03"}
    ))
    payload = jsonlib.loads(response.body)
    assert payload["generated_at"]
    assert [c["key"] for c in payload["chapters"]] == [
        "overview", "progress", "cost", "profit", "rating",
    ]


# ── report fallback domain rules ─────────────────────────────────────


def test_build_compare_report_handles_none_metrics():
    projects = [
        {"project_name": "甲", "contract": 100.0, "settlement": 0.0,
         "profit": 0.0, "progress": 0.0, "profit_rate": None,
         "unit_cost": None, "total_score": 44.0, "grade": "D"},
        {"project_name": "乙", "contract": 200.0, "settlement": 150.0,
         "profit": 30.0, "progress": 90.0, "profit_rate": 20.0,
         "unit_cost": 80.0, "total_score": 85.0, "grade": "A"},
    ]
    chapters = build_compare_report(projects, None)
    assert len(chapters) == 5
    assert all(isinstance(c["content"], str) and c["content"] for c in chapters)
    assert "乙" in chapters[4]["content"]
