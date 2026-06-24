import pytest
import openpyxl
from parser.core.cell_unmerger import unmerge


@pytest.fixture
def merged_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:A2")
    ws["A1"] = "merged_value"
    ws["B1"] = "normal_b1"
    ws["B2"] = "normal_b2"
    return wb


def test_unmerge_fills_merged_cells(merged_workbook):
    ws = merged_workbook.active
    grid = unmerge(ws)

    assert grid[0][0] == "merged_value"
    assert grid[0][1] == "normal_b1"
    assert grid[1][0] == "merged_value"
    assert grid[1][1] == "normal_b2"


def test_unmerge_preserves_empty_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "only_a1"
    ws["B1"] = ""  # force 2nd column
    grid = unmerge(ws)
    assert grid[0][0] == "only_a1"
    assert grid[0][1] == ""
