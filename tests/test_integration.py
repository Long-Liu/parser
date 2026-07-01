"""Integration tests for database init and template data tables.

Requires a running MySQL instance configured via config/local.yaml or env vars.
"""

import os

import openpyxl
import pytest
import sqlalchemy as sa

from core.pipeline import run_pipeline
from db.engine import get_sessionmaker
from utils.config_loader import list_configs, match_template


@pytest.fixture(scope="function")
def _config():
    """Load config; skip if not found."""
    try:
        from db.config import load_config
        return load_config()
    except FileNotFoundError:
        pytest.skip("Config file not found — integration tests require db setup")


@pytest.fixture(scope="function")
async def _db(_config):
    """Init module-level DB engine; skip if unavailable."""
    try:
        from db.engine import init as db_init, close as db_close
        await db_init(_config)
    except Exception:
        pytest.skip("Cannot connect to database")
    try:
        yield
    finally:
        await db_close()


@pytest.mark.asyncio
async def test_db_init(_config, _db):
    """Fixed tables are created successfully."""
    from db.schema import init_db

    await init_db()

    async with get_sessionmaker()() as session:
        result = await session.execute(
            sa.text("SELECT COUNT(*) AS cnt FROM information_schema.tables "
                    "WHERE table_schema=:db"),
            {"db": _config.DB_NAME},
        )
        count = result.scalar()
    assert count > 0


@pytest.mark.asyncio
async def test_data_tables_created(_db):
    """All template data tables are created."""
    from db.schema import create_data_table

    configs = list_configs()
    for cfg in configs:
        await create_data_table(cfg["template_id"])

    async with get_sessionmaker()() as session:
        for cfg in configs:
            result = await session.execute(
                sa.text("SHOW TABLES LIKE :tbl"),
                {"tbl": f"data_{cfg['template_id']}"}
            )
            row = result.first()
            assert row is not None, \
                f"Table data_{cfg['template_id']} not found"


@pytest.mark.asyncio
async def test_each_sheet_inserts_to_correct_table(_db):
    """Multi-sheet insert → each sheet's row count + data verified in the right table."""
    import os as _os

    from db.schema import init_db, create_data_table
    from services.data import insert_rows
    from repositories.data import DataRepo
    from repositories.batch import BatchRepo
    from repositories.project import ProjectRepo

    await init_db()
    for tid in ("labor_cost", "machinery"):
        await create_data_table(tid)

    proj = await ProjectRepo.get()
    if not proj:
        pid = await ProjectRepo.insert(code="TEST", name="test", created_by=1)
    else:
        pid = proj["id"]

    batch_id = await BatchRepo.insert(
        batch_no=f"TEST-{_os.urandom(4).hex()}", project_id=pid,
        ym="2025-06", uploaded_by=1, file_name="test.xlsx", file_size=100,
    )

    # Simulate 2 sheets: labor_cost (3 rows) + machinery (2 rows)
    sheets = {
        "labor_cost": [
            {"hierarchy_code": "1", "person_name": "Alice", "batch_id": batch_id},
            {"hierarchy_code": "2", "person_name": "Bob", "batch_id": batch_id},
            {"hierarchy_code": "3", "person_name": "Carol", "batch_id": batch_id},
        ],
        "machinery": [
            {"hierarchy_code": "A", "machine_name": "Crane", "batch_id": batch_id},
            {"hierarchy_code": "B", "machine_name": "Bulldozer", "batch_id": batch_id},
        ],
    }

    try:
        for template_id, rows in sheets.items():
            await insert_rows(template_id, rows)

        # Verify each sheet: count matches + data lands in correct table
        for template_id, expected_rows in sheets.items():
            result = await DataRepo.query(template_id, batch_id=batch_id)
            expected_count = len(expected_rows)
            assert result["total"] == expected_count, (
                f"[{template_id}] row count mismatch: "
                f"inserted {expected_count}, found {result['total']} in DB"
            )
            # Verify no cross-contamination: at least one value from each row is present
            for i, row in enumerate(expected_rows):
                db_row = result["rows"][i]
                for k, v in row.items():
                    if k == "batch_id":
                        continue
                    assert db_row.get(k) == v, (
                        f"[{template_id}] row {i}: expected {k}={v}, got {db_row.get(k)}"
                    )
    finally:
        for template_id in sheets:
            await DataRepo.delete(template_id, batch_id=batch_id)
        await BatchRepo.delete(BatchRepo._t().c.id == batch_id)


@pytest.mark.asyncio
async def test_real_excel_each_sheet_valid_rows_match_inserted_rows(_db):
    """For each real workbook sheet, parsed valid rows equal rows inserted into its table."""
    import uuid

    from db.schema import init_db, create_data_table
    from repositories.batch import BatchRepo
    from repositories.project import ProjectRepo
    from repositories.data import DataRepo
    from services.data import insert_rows

    excel_files = [
        os.path.join("excel", name)
        for name in os.listdir("excel")
        if name.endswith(".xlsx") and not name.startswith("~$")
    ]
    if not excel_files:
        pytest.skip("Excel file not found")

    await init_db()
    for cfg in list_configs():
        await create_data_table(cfg["template_id"])

    proj = await ProjectRepo.get()
    if not proj:
        pid = await ProjectRepo.insert(code=f"TEST-{uuid.uuid4().hex[:8]}", name="test", created_by=1)
    else:
        pid = proj["id"]

    batch_id = await BatchRepo.insert(
        batch_no=f"TEST-{uuid.uuid4().hex}", project_id=pid,
        ym="2025-06", uploaded_by=1, file_name=os.path.basename(excel_files[0]),
        file_size=os.path.getsize(excel_files[0]),
    )

    inserted_templates = set()
    try:
        wb = openpyxl.load_workbook(excel_files[0], data_only=True)
        for sheet_name in wb.sheetnames:
            config = match_template(sheet_name)
            if not config:
                continue

            result = run_pipeline(wb[sheet_name], batch_id=batch_id, config=config)
            template_id = result["template_id"]
            valid_row_count = len(result["rows"])
            assert result["success_rows"] == valid_row_count, (
                f"[{sheet_name}] success_rows={result['success_rows']} "
                f"but valid rows={valid_row_count}"
            )

            await insert_rows(template_id, result["rows"])
            inserted_templates.add(template_id)

            result_after_insert = await DataRepo.query(template_id, batch_id=batch_id)
            inserted_count = result_after_insert["total"]
            assert inserted_count == valid_row_count, (
                f"[{sheet_name} -> data_{template_id}] valid rows={valid_row_count}, "
                f"inserted rows={inserted_count}, total_rows={result['total_rows']}, "
                f"error_rows={result['error_rows']}, errors={result['errors'][:5]}"
            )
    finally:
        for template_id in inserted_templates:
            await DataRepo.delete(template_id, batch_id=batch_id)
        await BatchRepo.delete(BatchRepo._t().c.id == batch_id)


def test_full_parse_with_real_excel():
    """15-sheet full parse test."""
    excel_path = "excel/xxx项目主体施工动态成本表-样式 - 副本.xlsx"
    if not os.path.exists(excel_path):
        pytest.skip("Excel file not found")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    results = []
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        config = match_template(sheet_name)
        if config:
            result = run_pipeline(ws, batch_id=0, config=config)
            results.append(result)
            print(f"\n  {result['sheet_name']}: {result['success_rows']} rows "
                  f"({result['error_rows']} errors) [template: {result['template_id']}]")
        else:
            skipped += 1
            print(f"\n  {sheet_name}: SKIPPED")

    matched = len(results)
    total_rows = sum(r["success_rows"] for r in results)
    print(f"\nMatched: {matched}, Skipped: {skipped}, Total rows: {total_rows}")

    assert matched > 0, "At least one sheet should match"
    assert total_rows > 0, "Should extract some data"
