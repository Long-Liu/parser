"""GET /api/templates/<id>/download 与 xlsx 模板骨架生成测试。"""

import io
from types import SimpleNamespace
from urllib.parse import quote

import pytest
from openpyxl import load_workbook

from contexts.shared.domain.exceptions import NotFoundError
from contexts.template.application.template_app_service import (
    TemplateApplicationService,
)
from contexts.template.domain.template import (
    ColumnMapping,
    DynamicColumnMapping,
    HeaderSpec,
    HierarchyConfig,
    Template,
    TemplateId,
)
from contexts.template.infrastructure.repositories import YamlTemplateCatalog
from contexts.template.infrastructure.xlsx_template_builder import (
    build_template_workbook,
    sheet_name_for,
)
from contexts.template.infrastructure.yaml_loader import YamlTemplateLoader
from contexts.template.interface.template_controller import TemplatesController


# ── 骨架生成 ──────────────────────────────────────────────────────────

def test_sheet_name_strips_wildcards():
    template = Template(
        template_id=TemplateId("material_cost"),
        sheet_pattern="表9建安材料成本表*",
    )
    assert sheet_name_for(template) == "表9建安材料成本表"


def test_build_workbook_from_real_material_cost():
    template = YamlTemplateLoader().load("material_cost")
    content = build_template_workbook(template)
    assert content[:2] == b"PK"  # xlsx zip magic

    ws = load_workbook(io.BytesIO(content)).active
    assert ws.title == "表9建安材料成本表"

    # 两行表头：首列为层级序号列（落在最后一行表头）
    assert ws.cell(row=2, column=1).value == "序号"
    # fixed_columns：首个 match 关键词作为列名
    assert ws.cell(row=2, column=2).value == "成本科目"
    assert ws.cell(row=2, column=3).value == "单位"
    # 多关键词列：自上向下铺排
    assert ws.cell(row=1, column=4).value == "经济考核指标"
    assert ws.cell(row=2, column=4).value == "数量"
    # match_header 完全重复的列（已付/未付三组）只铺一次：
    # (首行, 末行) 表头组合不得重复
    header_pairs = [
        (ws.cell(row=1, column=c).value, ws.cell(row=2, column=c).value)
        for c in range(1, ws.max_column + 1)
    ]
    assert len(header_pairs) == len(set(header_pairs))
    # 数据起始行（第 3 行）留白
    assert all(
        ws.cell(row=3, column=c).value is None for c in range(1, ws.max_column + 1)
    )


def test_build_workbook_marks_dynamic_columns_with_example_month():
    template = Template(
        template_id=TemplateId("t"),
        sheet_pattern="样表*",
        header_spec=HeaderSpec(header_rows=[1, 2], data_start_row=3),
        hierarchy_config=HierarchyConfig(column_name="序号"),
        fixed_columns=[
            ColumnMapping(db_field="name", match_headers=["名称"]),
        ],
        dynamic_columns=[
            DynamicColumnMapping(db_prefix="monthly", match_headers=["2025年"]),
        ],
    )
    ws = load_workbook(io.BytesIO(build_template_workbook(template))).active
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["序号", "名称", "monthly_2026-01"]


# ── 端点 ──────────────────────────────────────────────────────────────
#
# handler 本体经 __wrapped__ 解包后以 fake request 测试；require_auth 通过
# isinstance(sanic.request.Request) 定位请求对象，因此 401 契约改为经真实
# Sanic app + ASGI 调用测试（同 tests/test_endpoint_smoke.py 风格）。

class _FakeAuthService:
    async def authenticate(self, token: str):
        if token != "good-token":
            from contexts.shared.domain.exceptions import AuthenticationError
            raise AuthenticationError("bad token")
        return SimpleNamespace(
            user_id=1, username="tester", permissions=set(),
            claims={"jti": "t", "iat": 0, "exp": 0},
        )


def _fake_request(token: str | None = "good-token"):
    headers = {"Authorization": f"Bearer {token}"} if token else {}
    return SimpleNamespace(
        headers=headers,
        ctx=SimpleNamespace(),
        app=SimpleNamespace(
            ctx=SimpleNamespace(
                services=SimpleNamespace(authorization=_FakeAuthService())
            )
        ),
    )


def _controller() -> TemplatesController:
    controller = TemplatesController(TemplateApplicationService(YamlTemplateCatalog()))
    controller.setup()
    return controller


async def test_download_endpoint_returns_xlsx_attachment():
    raw_handler = TemplatesController.download_template.__wrapped__
    response = await raw_handler(
        _controller(), _fake_request(), "material_cost"
    )
    assert response.status == 200
    assert "spreadsheetml.sheet" in response.content_type
    disposition = response.headers["Content-Disposition"]
    assert disposition.startswith("attachment;")
    assert 'filename="material_cost.xlsx"' in disposition
    assert f"filename*=UTF-8''{quote('建安材料成本表.xlsx')}" in disposition

    ws = load_workbook(io.BytesIO(response.body)).active
    assert ws.title == "表9建安材料成本表"
    assert ws.cell(row=2, column=2).value == "成本科目"


async def test_download_endpoint_unknown_template_404():
    raw_handler = TemplatesController.download_template.__wrapped__
    with pytest.raises(NotFoundError):
        await raw_handler(_controller(), _fake_request(), "no_such_tpl")


async def test_require_auth_decorator_returns_401_without_token():
    from sanic import Sanic
    from sanic.response import json

    from contexts.auth.interface.auth_middleware import require_auth

    app = Sanic("template_download_auth_smoke")
    app.asgi = True
    app.ctx.services = SimpleNamespace(authorization=_FakeAuthService())

    @app.get("/protected")
    @require_auth
    async def protected(request):
        return json({"ok": True})

    app.finalize()
    app.signalize(allow_fail_builtin=False)

    status: dict = {}

    async def receive():
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message):
        if message["type"] == "http.response.start":
            status["code"] = message["status"]

    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "GET",
        "scheme": "http",
        "path": "/protected",
        "raw_path": b"/protected",
        "query_string": b"",
        "root_path": "",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("127.0.0.1", 80),
    }
    await app(scope, receive, send)
    assert status.get("code") == 401


def test_download_route_registered_in_application():
    import application

    route_names = {route.name for route in application.app.router.routes}
    assert any(
        "template" in name and "download" in name for name in route_names
    )
