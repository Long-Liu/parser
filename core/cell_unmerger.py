"""Merged-cell expansion: fills merged-area values into all constituent cells."""

from openpyxl.worksheet.worksheet import Worksheet


def unmerge(worksheet: Worksheet) -> list[list]:
    """Expand merged cells so every cell contains the top-left value."""
    if not worksheet.max_row or not worksheet.max_column:
        return []

    max_row = worksheet.max_row if worksheet.max_row is not None else 1
    max_col = worksheet.max_column if worksheet.max_column is not None else 1

    grid = [[None] * max_col for _ in range(max_row)]

    for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), 1
    ):
        for col_idx, val in enumerate(row, 1):
            grid[row_idx - 1][col_idx - 1] = val

    for merged_range in worksheet.merged_cells.ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        top_left_value = grid[min_row - 1][min_col - 1]
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                grid[r - 1][c - 1] = top_left_value

    return grid
